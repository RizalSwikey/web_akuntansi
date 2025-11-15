# core/utils/excel_exporter.py
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- MODIFIED: Import both data calculators ---
from core.utils.final_report import (
    generate_final_report_data, 
    get_manufaktur_report_context
)

# --- STYLING (can be shared) ---
bold = Font(bold=True)
center = Alignment(horizontal="center")
right = Alignment(horizontal="right")
border_tb = Border(top=Side(style="thin"), bottom=Side(style="thin"))
header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
currency_format = "#,##0"


def _auto_fit_columns(ws):
    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = length + 3


def _generate_excel_dagang(wb, report, data):
    """
    Builds the Excel sheet for a "Dagang" report.
    This is your original function, now as a helper.
    """
    ws = wb.active
    ws.title = "Laporan Laba Rugi"

    # ========== HEADER ==========
    ws.append([report.company_name])
    ws.append(["LAPORAN LABA RUGI (DAGANG)"])
    ws.append([f"{report.month} {report.year}"])
    ws.append([])

    # ========== PENDAPATAN ==========
    ws.append(["Pendapatan"])
    ws.append(["Jenis", "Jumlah"])
    ws["A6"].font = bold
    ws["B6"].font = bold

    ws.append(["Pendapatan Usaha", data["total_pendapatan_usaha"]])
    ws.append(["Pendapatan Lain-lain", data["total_pendapatan_lain"]])
    ws.append(["Jumlah Pendapatan", data["jumlah_pendapatan"]])
    ws.append([])

    # ========== BEBAN ==========
    ws.append(["Beban"])
    ws.append(["Jenis", "Jumlah"])
    ws["A11"].font = bold
    ws["B11"].font = bold

    ws.append(["Harga Pokok Penjualan (HPP)", data["hpp_total"]])

    # Beban usaha & lainnya
    for item in data["beban_usaha_items"]:
        ws.append([item.name, item.total])
    ws.append(["Total Beban Usaha", data["total_beban_usaha"]])

    for item in data["beban_lain_items"]:
        ws.append([item.name, item.total])
    ws.append(["Total Beban Lain", data["total_beban_lain"]])

    ws.append(["Jumlah Beban", data["jumlah_beban"]])
    ws.append([])

    # ========== LABA ==========
    ws.append(["Laba/Rugi Sebelum Pajak", data["laba_sebelum_pajak"]])
    ws.append(["Beban Pajak Penghasilan", data["pajak_penghasilan"]])
    ws.append(["Laba/Rugi Setelah Pajak", data["laba_setelah_pajak"]])
    ws.append([])

    # ========== HPP PER PRODUK ==========
    ws.append(["Detail HPP Per Produk"])
    ws.append(["Produk", "HPP / Unit", "HPP Total"])
    
    header_row = ws.max_row
    for col in range(1, 4):
        ws.cell(row=header_row, column=col).font = bold

    for item in data["hpp_per_product"]:
        ws.append([
            item.get("product_name"),
            item.get("hpp_per_unit", 0),
            item.get("hpp", 0)
        ])
        
    _auto_fit_columns(ws)
    return wb


def _generate_excel_manufaktur(wb, report, data):
    """
    Builds the Excel sheets for a "Manufaktur" report.
    This is the new function.
    """
    
    ws_hpp = wb.active
    ws_hpp.title = "Lap. HPP"

    ws_hpp.append([report.company_name])
    ws_hpp.append(["LAPORAN HARGA POKOK PRODUKSI"])
    ws_hpp.append([f"{report.month} {report.year}"])
    ws_hpp.append([])

    ws_hpp.append(["Biaya Bahan Baku"])
    ws_hpp.cell(row=ws_hpp.max_row, column=1).font = bold
    ws_hpp.append(["Persediaan bahan baku (awal)", data['total_bb_awal']])
    ws_hpp.append(["Pembelian Bahan Baku", data['total_bb_pembelian']])
    ws_hpp.append(["Persediaan Bahan Baku (akhir)", -data['total_bb_akhir']])
    ws_hpp.append(["Total Biaya Bahan Baku", data['total_bbb']])
    ws_hpp.cell(row=ws_hpp.max_row, column=1).font = bold
    ws_hpp.append([])
    
    ws_hpp.append(["Biaya Tenaga Kerja Langsung", data['total_btkl']])
    ws_hpp.cell(row=ws_hpp.max_row, column=1).font = bold
    ws_hpp.append([])

    ws_hpp.append(["Biaya Overhead Pabrik"])
    ws_hpp.cell(row=ws_hpp.max_row, column=1).font = bold
    for item in data['bop_items']:
        ws_hpp.append([item.nama_biaya, item.total])
    ws_hpp.append(["Total Biaya Overhead Pabrik", data['total_bop']])
    ws_hpp.cell(row=ws_hpp.max_row, column=1).font = bold
    ws_hpp.append([])

    ws_hpp.append(["Total Biaya Produksi", data['total_biaya_produksi']])
    ws_hpp.cell(row=ws_hpp.max_row, column=1).font = bold
    ws_hpp.append(["Persediaan BDP (Awal)", data['total_bdp_awal']])
    ws_hpp.append(["Persediaan BDP (Akhir)", -data['total_bdp_akhir']])
    ws_hpp.append(["Cost of Goods Manufactured (COGM)", data['cogm']])
    ws_hpp.cell(row=ws_hpp.max_row, column=1).font = bold
    ws_hpp.cell(row=ws_hpp.max_row, column=1).fill = total_fill
    ws_hpp.cell(row=ws_hpp.max_row, column=2).fill = total_fill
    ws_hpp.append([])
    
    ws_hpp.append(["Persediaan Barang Jadi (Awal)", data['total_bj_awal']])
    ws_hpp.append(["Barang Siap untuk Dijual", data['barang_siap_dijual']])
    ws_hpp.append(["Persediaan Barang Jadi (Akhir)", -data['total_bj_akhir']])
    ws_hpp.append(["Harga Pokok Penjualan (HPP)", data['hpp_total']])
    ws_hpp.cell(row=ws_hpp.max_row, column=1).font = bold
    ws_hpp.cell(row=ws_hpp.max_row, column=1).fill = total_fill
    ws_hpp.cell(row=ws_hpp.max_row, column=2).fill = total_fill
    ws_hpp.append([])

    _auto_fit_columns(ws_hpp)

    ws_lr = wb.create_sheet(title="Lap. Laba Rugi")
    
    ws_lr.append([report.company_name])
    ws_lr.append(["LAPORAN LABA RUGI (MANUFAKTUR)"])
    ws_lr.append([f"{report.month} {report.year}"])
    ws_lr.append([])
    
    ws_lr.append(["Pendapatan"])
    ws_lr.cell(row=ws_lr.max_row, column=1).font = bold
    ws_lr.append(["Pendapatan Usaha", data['total_pendapatan_usaha']])
    ws_lr.append(["Pendapatan Lain-lain", data['total_pendapatan_lain']])
    ws_lr.append(["Jumlah Pendapatan", data['jumlah_pendapatan']])
    ws_lr.cell(row=ws_lr.max_row, column=1).font = bold
    ws_lr.append([])
    
    ws_lr.append(["Beban"])
    ws_lr.cell(row=ws_lr.max_row, column=1).font = bold
    ws_lr.append(["Harga Pokok Penjualan (HPP)", data['hpp_total']])
    for item in data['beban_usaha_items']:
        ws_lr.append([item.name, item.total])
    ws_lr.append(["Total Beban Usaha Lainnya", data['total_beban_usaha_lainnya']])
    for item in data['beban_lain_items']:
        ws_lr.append([item.name, item.total])
    ws_lr.append(["Total Beban Lain-lain", data['total_beban_lain']])
    ws_lr.append(["Jumlah Beban", data['jumlah_beban']])
    ws_lr.cell(row=ws_lr.max_row, column=1).font = bold
    ws_lr.append([])
    
    ws_lr.append(["Laba/Rugi Sebelum Pajak", data['laba_sebelum_pajak']])
    ws_lr.append(["Beban Pajak Penghasilan", data['pajak_penghasilan']])
    ws_lr.append(["Laba/Rugi Setelah Pajak", data['laba_setelah_pajak']])
    ws_lr.cell(row=ws_lr.max_row, column=1).font = bold
    ws_lr.cell(row=ws_lr.max_row, column=1).fill = total_fill
    ws_lr.cell(row=ws_lr.max_row, column=2).fill = total_fill
    
    _auto_fit_columns(ws_lr)
    
    return wb


def generate_excel_file(report):
    
    wb = Workbook()

    if report.business_type == 'manufaktur':
        data = get_manufaktur_report_context(report)
        wb = _generate_excel_manufaktur(wb, report, data)
        filename = f"Laporan_Manufaktur_{report.company_name}_{report.month}_{report.year}.xlsx"
    else:
        # Default to Dagang
        data = generate_final_report_data(report)
        wb = _generate_excel_dagang(wb, report, data)
        filename = f"Laporan_Dagang_{report.company_name}_{report.month}_{report.year}.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue(), filename